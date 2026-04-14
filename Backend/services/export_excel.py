"""Excel invoice export using openpyxl — IPC_Invoice_System_v4 structure."""
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# Brand colours (ARGB: FF + hex)
BLUE = "FF1e3a5f"
LIGHT_BLUE = "FFe8f0fe"
GREY = "FF6b7280"
LIGHT_GREY = "FFf9fafb"
WHITE = "FFFFFFFF"
RED = "FFdc2626"
HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=10)
BLUE_FILL = PatternFill("solid", fgColor=BLUE)
LIGHT_BLUE_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
STRIPE_FILL = PatternFill("solid", fgColor=LIGHT_GREY)
LABEL_FONT = Font(name="Calibri", bold=True, size=9, color=BLUE)
BODY_FONT = Font(name="Calibri", size=9)
GREY_FONT = Font(name="Calibri", size=9, color=GREY)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=BLUE)
TOTAL_FONT = Font(name="Calibri", bold=True, size=11, color=BLUE)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top")

MONEY_FMT = '#,##0.00'
DATE_FMT = 'YYYY-MM-DD'


def _thin_border(sides="all"):
    thin = Side(style="thin", color="FFd1d5db")
    b = Border()
    if "all" in sides or "left" in sides:
        b.left = thin
    if "all" in sides or "right" in sides:
        b.right = thin
    if "all" in sides or "top" in sides:
        b.top = thin
    if "all" in sides or "bottom" in sides:
        b.bottom = thin
    return b


def _set_col_widths(ws, widths: dict[str, float]):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _header_row(ws, row: int, headers: list[str], fill=BLUE_FILL, font=HEADER_FONT):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = _thin_border()


def _data_row(ws, row: int, values: list[Any], stripe: bool = False):
    fill = STRIPE_FILL if stripe else PatternFill("solid", fgColor=WHITE)
    border = _thin_border()
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = BODY_FONT
        cell.fill = fill
        cell.border = border
        cell.alignment = LEFT
        if isinstance(val, float) and col_idx > 1:
            cell.number_format = MONEY_FMT
            cell.alignment = RIGHT


def _money(val) -> float:
    return round(float(val or 0), 2)


def generate_invoice_xlsx(edit_data: dict) -> bytes:
    """
    Generate a .xlsx export for an invoice. Returns raw bytes.
    """
    invoice = edit_data.get("invoice", {})
    client = edit_data.get("client") or {}
    project = edit_data.get("project") or {}
    lines = edit_data.get("lines", [])
    expenses = edit_data.get("expenses", [])

    inv_number = invoice.get("invoice_number") or invoice.get("id", "")[:8]
    inv_label = f"INV-{inv_number}"
    status = invoice.get("status") or "draft"
    cap_amount = invoice.get("cap_amount")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ─────────────────────────── Sheet: Invoice ──────────────────────────────
    ws_inv = wb.create_sheet("Invoice")
    ws_inv.sheet_view.showGridLines = False

    # Title block
    ws_inv.merge_cells("A1:G1")
    c = ws_inv["A1"]
    c.value = "Impact Point Co., LLC"
    c.font = TITLE_FONT
    c.alignment = LEFT

    ws_inv.merge_cells("A2:G2")
    c = ws_inv["A2"]
    c.value = inv_label
    c.font = Font(name="Calibri", bold=True, size=12, color=BLUE)
    c.alignment = LEFT

    # Meta fields
    meta = [
        ("Status:", status.upper()),
        ("Issue Date:", invoice.get("issue_date")),
        ("Due Date:", invoice.get("due_date")),
        ("Client:", client.get("name", "")),
        ("Project:", project.get("name", "")),
        ("Notes:", invoice.get("notes", "")),
    ]
    for i, (label, val) in enumerate(meta, 4):
        ws_inv.cell(row=i, column=1, value=label).font = LABEL_FONT
        c = ws_inv.cell(row=i, column=2, value=val)
        c.font = BODY_FONT
        c.alignment = LEFT

    ws_inv.column_dimensions["A"].width = 14
    ws_inv.column_dimensions["B"].width = 28
    ws_inv.row_dimensions[1].height = 22
    ws_inv.row_dimensions[2].height = 18

    # ─────────────────────────── Sheet: Time_Detail ──────────────────────────
    ws_time = wb.create_sheet("Time_Detail")
    ws_time.sheet_view.showGridLines = False

    headers_time = [
        "Invoice", "Employee", "Title", "Role",
        "Hours", "Rate (USD)", "Subtotal", "Disc. Type", "Disc. Value",
        "Disc. ($)", "Total"
    ]
    _header_row(ws_time, 1, headers_time)

    ROLE_LABELS = {
        "manager": "Hours Cost per Employee",
        "contractor": "Professional Fees",
        "employee": "Billable Hours",
    }

    total_fees = 0.0
    total_discounts = 0.0
    for i, line in enumerate(lines, 2):
        hours = _money(line.get("hours", 0))
        rate = _money(line.get("hourly_rate", 0))
        subtotal = hours * rate
        disc_type = line.get("discount_type") or "amount"
        disc_val = _money(line.get("discount_value", 0))
        disc_dollars = (subtotal * disc_val / 100) if disc_type == "percent" else disc_val
        line_total = max(0.0, subtotal - disc_dollars)
        total_fees += line_total
        total_discounts += disc_dollars

        role_key = (line.get("role") or "employee").lower()
        _data_row(ws_time, i, [
            inv_label,
            line.get("employee_name", ""),
            line.get("title", ""),
            ROLE_LABELS.get(role_key, role_key.title()),
            hours,
            rate,
            subtotal,
            disc_type,
            disc_val,
            disc_dollars,
            line_total,
        ], stripe=i % 2 == 0)

    # Totals row
    totals_row = len(lines) + 2
    ws_time.cell(row=totals_row, column=1, value="TOTALS").font = LABEL_FONT
    for col, val in [(7, total_fees + total_discounts), (10, total_discounts), (11, total_fees)]:
        c = ws_time.cell(row=totals_row, column=col, value=_money(val))
        c.font = TOTAL_FONT
        c.number_format = MONEY_FMT
        c.alignment = RIGHT
        c.fill = LIGHT_BLUE_FILL

    _set_col_widths(ws_time, {
        "A": 16, "B": 22, "C": 20, "D": 22,
        "E": 8, "F": 11, "G": 11, "H": 10,
        "I": 10, "J": 11, "K": 11,
    })

    # ─────────────────────────── Sheet: Expense_Details ──────────────────────
    ws_exp = wb.create_sheet("Expense_Details")
    ws_exp.sheet_view.showGridLines = False

    headers_exp = [
        "Invoice", "Date", "Professional", "Vendor",
        "Description", "Category", "Amount (USD)",
        "Payment Source", "Receipt Attached", "Notes"
    ]
    _header_row(ws_exp, 1, headers_exp)

    total_expenses = 0.0
    for i, exp in enumerate(expenses, 2):
        amount = _money(exp.get("amount_usd", 0))
        total_expenses += amount
        _data_row(ws_exp, i, [
            inv_label,
            exp.get("date"),
            exp.get("professional") or "",
            exp.get("vendor") or "",
            exp.get("description") or "",
            exp.get("category") or "",
            amount,
            exp.get("payment_source") or "",
            "Yes" if exp.get("receipt_attached") else "No",
            exp.get("notes") or "",
        ], stripe=i % 2 == 0)

    # Expense totals
    exp_totals_row = len(expenses) + 2
    ws_exp.cell(row=exp_totals_row, column=1, value="TOTAL").font = LABEL_FONT
    c = ws_exp.cell(row=exp_totals_row, column=7, value=_money(total_expenses))
    c.font = TOTAL_FONT
    c.number_format = MONEY_FMT
    c.alignment = RIGHT
    c.fill = LIGHT_BLUE_FILL

    # Category subtotals
    cat_totals: dict[str, float] = {}
    for exp in expenses:
        cat = exp.get("category") or "Other"
        cat_totals[cat] = cat_totals.get(cat, 0) + _money(exp.get("amount_usd", 0))
    sub_row = exp_totals_row + 2
    ws_exp.cell(row=sub_row, column=1, value="By Category:").font = LABEL_FONT
    for j, (cat, val) in enumerate(cat_totals.items(), 2):
        ws_exp.cell(row=sub_row, column=j, value=f"{cat}: ${val:,.2f}").font = GREY_FONT

    _set_col_widths(ws_exp, {
        "A": 16, "B": 12, "C": 18, "D": 18,
        "E": 26, "F": 20, "G": 13, "H": 16,
        "I": 16, "J": 22,
    })

    # ─────────────────────────── Sheet: Summary ──────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.sheet_view.showGridLines = False

    ws_sum.merge_cells("A1:C1")
    c = ws_sum["A1"]
    c.value = f"Invoice Summary — {inv_label}"
    c.font = TITLE_FONT

    summary_data = [
        ("Invoice Number", inv_label),
        ("Status", status.upper()),
        ("Client", client.get("name", "")),
        ("Project", project.get("name", "")),
        ("Issue Date", str(invoice.get("issue_date") or "")),
        ("Due Date", str(invoice.get("due_date") or "")),
        (None, None),
        ("Total Fees (gross)", _money(total_fees + total_discounts)),
        ("Total Discounts", _money(total_discounts)),
        ("Subtotal Fees (net)", _money(total_fees)),
    ]
    if cap_amount is not None:
        summary_data.append(("Cap Amount", _money(cap_amount)))
        capped = min(total_fees, _money(cap_amount))
    else:
        capped = total_fees

    summary_data.append(("Total Expenses", _money(total_expenses)))
    summary_data.append((None, None))
    summary_data.append(("TOTAL DUE", _money(capped + total_expenses)))

    for i, (label, val) in enumerate(summary_data, 3):
        if label is None:
            continue
        c_label = ws_sum.cell(row=i, column=1, value=label)
        c_val = ws_sum.cell(row=i, column=2, value=val)
        if label in ("TOTAL DUE",):
            c_label.font = TOTAL_FONT
            c_val.font = TOTAL_FONT
            c_val.number_format = MONEY_FMT
            c_label.fill = LIGHT_BLUE_FILL
            c_val.fill = LIGHT_BLUE_FILL
        elif isinstance(val, float):
            c_label.font = BODY_FONT
            c_val.font = BODY_FONT
            c_val.number_format = MONEY_FMT
            c_val.alignment = RIGHT
        else:
            c_label.font = LABEL_FONT
            c_val.font = BODY_FONT

    _set_col_widths(ws_sum, {"A": 26, "B": 18, "C": 10})
    ws_sum.row_dimensions[1].height = 22

    # ─────────────────────────── Sheet: Clients ──────────────────────────────
    ws_cli = wb.create_sheet("Clients")
    ws_cli.sheet_view.showGridLines = False
    _header_row(ws_cli, 1, ["Client Name", "Email", "Phone", "Project"])
    _data_row(ws_cli, 2, [
        client.get("name", ""),
        client.get("email", ""),
        client.get("phone", ""),
        project.get("name", ""),
    ])
    _set_col_widths(ws_cli, {"A": 24, "B": 26, "C": 16, "D": 24})

    # ─────────────────────────── Sheet: Professionals ────────────────────────
    ws_pro = wb.create_sheet("Professionals")
    ws_pro.sheet_view.showGridLines = False
    _header_row(ws_pro, 1, ["Employee", "Title", "Role", "Hourly Rate", "Total Hours", "Net Total"])

    seen: dict[str, dict] = {}
    for line in lines:
        uid = line.get("user_id") or line.get("employee_name")
        if uid not in seen:
            seen[uid] = {
                "name": line.get("employee_name", ""),
                "title": line.get("title", ""),
                "role": line.get("role", ""),
                "rate": _money(line.get("hourly_rate", 0)),
                "hours": 0.0,
                "total": 0.0,
            }
        seen[uid]["hours"] += _money(line.get("hours", 0))
        hours = _money(line.get("hours", 0))
        rate = _money(line.get("hourly_rate", 0))
        subtotal = hours * rate
        disc_type = line.get("discount_type") or "amount"
        disc_val = _money(line.get("discount_value", 0))
        disc_dollars = (subtotal * disc_val / 100) if disc_type == "percent" else disc_val
        seen[uid]["total"] += max(0.0, subtotal - disc_dollars)

    for i, pro in enumerate(seen.values(), 2):
        _data_row(ws_pro, i, [
            pro["name"], pro["title"], pro["role"],
            pro["rate"], pro["hours"], pro["total"],
        ], stripe=i % 2 == 0)

    _set_col_widths(ws_pro, {"A": 22, "B": 24, "C": 18, "D": 13, "E": 13, "F": 13})

    # ─────────────────────────── Sheet: Config ───────────────────────────────
    ws_cfg = wb.create_sheet("Config")
    ws_cfg.sheet_view.showGridLines = False
    _header_row(ws_cfg, 1, ["Key", "Value"])
    config_rows = [
        ("Company", "Impact Point Co., LLC"),
        ("Export Tool", "H_TRACKER v1.0"),
        ("Invoice", inv_label),
        ("Generated", str(__import__("datetime").date.today())),
    ]
    for i, (k, v) in enumerate(config_rows, 2):
        _data_row(ws_cfg, i, [k, v])
    _set_col_widths(ws_cfg, {"A": 20, "B": 30})

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_invoices_report_xlsx(invoices_data: list) -> bytes:
    """
    Generate a multi-invoice report .xlsx. Returns raw bytes.
    invoices_data: list of dicts with keys invoice, client, project, lines, expenses.

    Sheet layout:
      1. Invoice_Detail  — full block per invoice (header + lines + expenses + totals)
      2. Time_Detail     — flat table of all lines across all invoices
      3. Expenses        — flat table of all expenses across all invoices
    """
    import datetime as dt

    # ── shared styles ──────────────────────────────────────────────────────────
    INV_TITLE_FILL = PatternFill("solid", fgColor="FF1e3a5f")   # dark blue
    INV_TITLE_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFFFF")
    META_FILL      = PatternFill("solid", fgColor="FFe8f0fe")    # light blue
    META_FONT      = Font(name="Calibri", bold=True, size=9, color="FF1e3a5f")
    META_VAL_FONT  = Font(name="Calibri", size=9, color="FF374151")
    SEC_FILL       = PatternFill("solid", fgColor="FFf3f4f6")    # light grey for sub-headers
    SEC_FONT       = Font(name="Calibri", bold=True, size=9, color="FF6b7280")
    TOTAL_FILL     = PatternFill("solid", fgColor="FFdbeafe")    # pale blue for invoice total
    GRAND_FILL     = PatternFill("solid", fgColor="FF1e3a5f")
    GRAND_FONT     = Font(name="Calibri", bold=True, size=11, color="FFFFFFFF")

    NCOLS = 10  # A–J

    def _merge_title(ws, row, value, fill, font, height=18):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
        c = ws.cell(row=row, column=1, value=value)
        c.font = font
        c.fill = fill
        c.alignment = LEFT
        c.border = _thin_border()
        ws.row_dimensions[row].height = height
        # apply fill+border to merged tail cells too
        for col in range(2, NCOLS + 1):
            tc = ws.cell(row=row, column=col)
            tc.fill = fill
            tc.border = _thin_border()

    def _meta_pair(ws, row, pairs: list):
        """Write alternating label/value pairs across the row (up to 5 pairs = 10 cols)."""
        for idx, (label, val) in enumerate(pairs):
            col = idx * 2 + 1
            lc = ws.cell(row=row, column=col, value=label)
            lc.font = META_FONT
            lc.fill = META_FILL
            lc.alignment = LEFT
            lc.border = _thin_border()
            vc = ws.cell(row=row, column=col + 1, value=val)
            vc.font = META_VAL_FONT
            vc.fill = META_FILL
            vc.alignment = LEFT
            vc.border = _thin_border()
        # fill remaining columns
        filled = len(pairs) * 2
        for col in range(filled + 1, NCOLS + 1):
            tc = ws.cell(row=row, column=col)
            tc.fill = META_FILL
            tc.border = _thin_border()

    def _section_header(ws, row, headers: list, ncols_override=None):
        n = ncols_override or NCOLS
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = SEC_FONT
            c.fill = SEC_FILL
            c.alignment = CENTER
            c.border = _thin_border()
        for col in range(len(headers) + 1, n + 1):
            ws.cell(row=row, column=col).fill = SEC_FILL

    def _blank_row(ws, row):
        for col in range(1, NCOLS + 1):
            ws.cell(row=row, column=col).value = None

    def _subtotal_row(ws, row, label, value, label_col=1, value_col=10):
        for col in range(1, NCOLS + 1):
            c = ws.cell(row=row, column=col)
            c.fill = TOTAL_FILL
            c.border = _thin_border()
        lc = ws.cell(row=row, column=label_col, value=label)
        lc.font = LABEL_FONT
        lc.fill = TOTAL_FILL
        vc = ws.cell(row=row, column=value_col, value=_money(value))
        vc.font = TOTAL_FONT
        vc.number_format = MONEY_FMT
        vc.alignment = RIGHT
        vc.fill = TOTAL_FILL

    wb = Workbook()
    wb.remove(wb.active)

    # ═══════════════════════ Sheet 1: Invoice_Detail ══════════════════════════
    ws = wb.create_sheet("Invoice_Detail")
    ws.sheet_view.showGridLines = False

    # Report title
    today_str = dt.date.today().strftime("%B %d, %Y")
    _merge_title(ws, 1, f"INVOICES REPORT  ·  Impact Point Co., LLC  ·  {today_str}",
                 INV_TITLE_FILL, INV_TITLE_FONT, height=24)

    cur_row = 3  # start after title + blank

    grand_fees = 0.0
    grand_exp  = 0.0
    grand_tot  = 0.0

    for data in invoices_data:
        inv     = data.get("invoice", {})
        client  = data.get("client") or {}
        project = data.get("project") or {}
        lines   = data.get("lines", [])
        expenses = data.get("expenses", [])

        inv_num  = inv.get("invoice_number") or inv.get("id", "")[:8]
        company  = inv.get("owner_company") or "IPC"
        status   = (inv.get("status") or "draft").upper()
        proj_name = project.get("name", "—")
        cli_name  = client.get("name", "—")

        # ── Invoice title bar ─────────────────────────────────────────────
        _merge_title(ws, cur_row,
                     f"  #{inv_num}   {proj_name}   |   {cli_name}   [{company}]",
                     INV_TITLE_FILL, INV_TITLE_FONT, height=20)
        cur_row += 1

        # ── Meta row 1: status, period, issue, due ────────────────────────
        _meta_pair(ws, cur_row, [
            ("Status",       status),
            ("Period",       f"{inv.get('period_start') or '—'}  →  {inv.get('period_end') or '—'}"),
            ("Issue Date",   str(inv.get("issue_date") or "—")),
            ("Due Date",     str(inv.get("due_date") or "—")),
            ("Cap Amount",   f"${_money(inv.get('cap_amount', 0)):,.2f}" if inv.get("cap_amount") else "—"),
        ])
        cur_row += 1

        # ── Meta row 2: client contact, notes ─────────────────────────────
        notes_val = (inv.get("notes") or "")[:60] or "—"
        _meta_pair(ws, cur_row, [
            ("Client Email",   client.get("email") or "—"),
            ("Client Phone",   client.get("phone") or "—"),
            ("Signatory",      inv.get("signatory_name") or "—"),
            ("Notes",          notes_val),
        ])
        cur_row += 1

        # ── Lines sub-table ───────────────────────────────────────────────
        _section_header(ws, cur_row,
                        ["Employee", "Title", "Hours", "Rate (USD)",
                         "Subtotal", "Disc. Type", "Disc. Value", "Disc. ($)", "Net Total", ""])
        cur_row += 1

        lines_net = 0.0
        for line in lines:
            hours     = _money(line.get("hours", 0))
            rate      = _money(line.get("hourly_rate", 0))
            subtotal  = hours * rate
            disc_type = line.get("discount_type") or "amount"
            disc_val  = _money(line.get("discount_value", 0))
            disc_dol  = (subtotal * disc_val / 100) if disc_type == "percent" else disc_val
            net       = max(0.0, subtotal - disc_dol)
            lines_net += net
            stripe = cur_row % 2 == 0
            fill   = STRIPE_FILL if stripe else PatternFill("solid", fgColor=WHITE)
            vals   = [
                line.get("employee_name", ""),
                line.get("title", ""),
                hours, rate, subtotal,
                disc_type, disc_val, disc_dol, net, "",
            ]
            for col_idx, val in enumerate(vals, 1):
                c = ws.cell(row=cur_row, column=col_idx, value=val)
                c.font  = BODY_FONT
                c.fill  = fill
                c.border = _thin_border()
                c.alignment = RIGHT if isinstance(val, float) else LEFT
                if isinstance(val, float):
                    c.number_format = MONEY_FMT
            cur_row += 1

        # Lines subtotal
        _subtotal_row(ws, cur_row, "Lines Net Total", lines_net)
        cur_row += 1
        grand_fees += lines_net

        # ── Expenses sub-table (only if present) ──────────────────────────
        exp_total = 0.0
        if expenses:
            _blank_row(ws, cur_row)
            cur_row += 1
            _section_header(ws, cur_row,
                            ["Date", "Professional", "Vendor", "Description",
                             "Category", "Amount (USD)", "Payment Source", "Receipt", "Notes", ""])
            cur_row += 1
            for exp in expenses:
                amount = _money(exp.get("amount_usd", 0))
                exp_total += amount
                stripe = cur_row % 2 == 0
                fill   = STRIPE_FILL if stripe else PatternFill("solid", fgColor=WHITE)
                vals   = [
                    str(exp.get("date") or ""),
                    exp.get("professional") or "",
                    exp.get("vendor") or "",
                    exp.get("description") or "",
                    exp.get("category") or "",
                    amount,
                    exp.get("payment_source") or "",
                    "Yes" if exp.get("receipt_attached") else "No",
                    exp.get("notes") or "",
                    "",
                ]
                for col_idx, val in enumerate(vals, 1):
                    c = ws.cell(row=cur_row, column=col_idx, value=val)
                    c.font   = BODY_FONT
                    c.fill   = fill
                    c.border = _thin_border()
                    c.alignment = RIGHT if isinstance(val, float) else LEFT
                    if isinstance(val, float):
                        c.number_format = MONEY_FMT
                cur_row += 1
            _subtotal_row(ws, cur_row, "Expenses Total", exp_total)
            cur_row += 1
            grand_exp += exp_total

        # ── Invoice grand total ────────────────────────────────────────────
        inv_total = _money(inv.get("total", 0)) + exp_total
        grand_tot += inv_total
        for col in range(1, NCOLS + 1):
            c = ws.cell(row=cur_row, column=col)
            c.fill   = INV_TITLE_FILL
            c.border = _thin_border()
        lc = ws.cell(row=cur_row, column=1, value=f"INVOICE TOTAL  #{inv_num}")
        lc.font = INV_TITLE_FONT
        lc.fill = INV_TITLE_FILL
        vc = ws.cell(row=cur_row, column=NCOLS, value=_money(inv_total))
        vc.font = INV_TITLE_FONT
        vc.number_format = MONEY_FMT
        vc.alignment = RIGHT
        vc.fill = INV_TITLE_FILL
        cur_row += 1

        # ── Spacer between invoices ────────────────────────────────────────
        _blank_row(ws, cur_row)
        cur_row += 1
        _blank_row(ws, cur_row)
        cur_row += 1

    # ── Grand total row ────────────────────────────────────────────────────
    for col in range(1, NCOLS + 1):
        c = ws.cell(row=cur_row, column=col)
        c.fill   = GRAND_FILL
        c.border = _thin_border()
    ws.row_dimensions[cur_row].height = 20
    lc = ws.cell(row=cur_row, column=1,
                 value=f"GRAND TOTAL  ({len(invoices_data)} invoices)")
    lc.font = GRAND_FONT
    lc.fill = GRAND_FILL
    for col, label, val in [
        (6, "Fees",     grand_fees),
        (8, "Expenses", grand_exp),
        (10, "TOTAL",   grand_tot),
    ]:
        hc = ws.cell(row=cur_row, column=col - 1, value=label)
        hc.font = Font(name="Calibri", bold=True, size=9, color="FFbfdbfe")
        hc.fill = GRAND_FILL
        hc.alignment = RIGHT
        vc = ws.cell(row=cur_row, column=col, value=_money(val))
        vc.font = GRAND_FONT
        vc.fill = GRAND_FILL
        vc.number_format = MONEY_FMT
        vc.alignment = RIGHT

    _set_col_widths(ws, {
        "A": 22, "B": 22, "C": 9, "D": 11,
        "E": 11, "F": 11, "G": 11, "H": 11,
        "I": 11, "J": 11,
    })
    ws.freeze_panes = "A2"

    # ═══════════════════════ Sheet 2: Time_Detail ═════════════════════════════
    ws_time = wb.create_sheet("Time_Detail")
    ws_time.sheet_view.showGridLines = False

    headers_time = [
        "Invoice #", "Company", "Project", "Employee", "Title",
        "Hours", "Rate (USD)", "Subtotal", "Disc. Type", "Disc. Value",
        "Disc. ($)", "Total",
    ]
    _header_row(ws_time, 1, headers_time)

    row = 2
    for data in invoices_data:
        inv     = data.get("invoice", {})
        project = data.get("project") or {}
        inv_num = inv.get("invoice_number") or inv.get("id", "")[:8]
        company = inv.get("owner_company") or "IPC"
        for line in data.get("lines", []):
            hours     = _money(line.get("hours", 0))
            rate      = _money(line.get("hourly_rate", 0))
            subtotal  = hours * rate
            disc_type = line.get("discount_type") or "amount"
            disc_val  = _money(line.get("discount_value", 0))
            disc_dol  = (subtotal * disc_val / 100) if disc_type == "percent" else disc_val
            net       = max(0.0, subtotal - disc_dol)
            _data_row(ws_time, row, [
                f"#{inv_num}", company, project.get("name", ""),
                line.get("employee_name", ""), line.get("title", ""),
                hours, rate, subtotal, disc_type, disc_val, disc_dol, net,
            ], stripe=row % 2 == 0)
            row += 1

    _set_col_widths(ws_time, {
        "A": 14, "B": 8, "C": 24, "D": 22, "E": 20,
        "F": 8,  "G": 11, "H": 11, "I": 10, "J": 10,
        "K": 11, "L": 11,
    })

    # ═══════════════════════ Sheet 3: Expenses ════════════════════════════════
    ws_exp = wb.create_sheet("Expenses")
    ws_exp.sheet_view.showGridLines = False

    headers_exp = [
        "Invoice #", "Company", "Project", "Date", "Professional",
        "Vendor", "Description", "Category", "Amount (USD)",
        "Payment Source", "Receipt", "Notes",
    ]
    _header_row(ws_exp, 1, headers_exp)

    row = 2
    for data in invoices_data:
        inv     = data.get("invoice", {})
        project = data.get("project") or {}
        inv_num = inv.get("invoice_number") or inv.get("id", "")[:8]
        company = inv.get("owner_company") or "IPC"
        for exp in data.get("expenses", []):
            _data_row(ws_exp, row, [
                f"#{inv_num}", company, project.get("name", ""),
                str(exp.get("date") or ""),
                exp.get("professional") or "",
                exp.get("vendor") or "",
                exp.get("description") or "",
                exp.get("category") or "",
                _money(exp.get("amount_usd", 0)),
                exp.get("payment_source") or "",
                "Yes" if exp.get("receipt_attached") else "No",
                exp.get("notes") or "",
            ], stripe=row % 2 == 0)
            row += 1

    _set_col_widths(ws_exp, {
        "A": 14, "B": 8, "C": 22, "D": 12, "E": 18,
        "F": 18, "G": 26, "H": 18, "I": 13,
        "J": 16, "K": 10, "L": 22,
    })

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
